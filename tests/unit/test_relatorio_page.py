"""
Testes para funções auxiliares de pages/01_relatorio.py.
A lógica de UI (AgGrid, st.button) não é testável unitariamente — está coberta
pelos testes de report_service.py e factor_service.py. Aqui validamos:
- Conversão de ReportRow → DataFrame (ordem, rótulos, valores None)
- Exportação Excel (estrutura, tipos numéricos, arredondamento)
"""
import sys
import io
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

# Importar apenas funções puras da página — sem inicializar Streamlit
from models.report_row import CorDiferenca, ReportRow


def _make_row(
    data="2026-03-15",
    artigo="Artigo A",
    cor="PRETO",
    lote="L001",
    codigo="ART001",
    m2=100.0,
    peso_lote=120.0,
    kg_m2=1.20,
    kg_ft2=0.11,
    fator=1.25,
    peso_calc=124.50,
    pct=-3.62,
    cor_dif=CorDiferenca.VERDE,
) -> ReportRow:
    return ReportRow(
        data_recurtimento=data,
        artigo=artigo,
        cor=cor,
        lote_fabricacao=lote,
        codigo_artigo=codigo,
        m2=m2,
        peso_lote=peso_lote,
        kg_m2=kg_m2,
        kg_ft2=kg_ft2,
        fator_aplicado=fator,
        peso_calculado=peso_calc,
        pct_diferenca=pct,
        cor_diferenca=cor_dif,
    )


def _make_row_sem_calculo() -> ReportRow:
    return ReportRow(
        data_recurtimento="2026-03-15",
        artigo="Artigo B",
        cor="MARROM",
        lote_fabricacao="L002",
        codigo_artigo="ART999",
        m2=50.0,
        peso_lote=None,
    )


# Importar funções puras sem acionar st.set_page_config
def _import_page_functions():
    """Importa apenas as funções auxiliares sem executar o bloco Streamlit."""
    import importlib, types, unittest.mock as mock

    # Mock streamlit para evitar erros de contexto
    st_mock = mock.MagicMock()
    st_mock.cache_resource = lambda f: f  # passa a função sem decorar
    sys.modules["streamlit"] = st_mock
    sys.modules["st_aggrid"] = mock.MagicMock()

    # Importar o módulo da página com mocks ativos
    spec = importlib.util.spec_from_file_location(
        "relatorio_page",
        Path(__file__).parent.parent.parent / "pages" / "01_relatorio.py",
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture(scope="module")
def page():
    return _import_page_functions()


# ---------------------------------------------------------------------------
# _rows_to_dataframe
# ---------------------------------------------------------------------------

class TestRowsToDataframe:
    def test_colunas_na_ordem_correta(self, page):
        df = page._rows_to_dataframe([_make_row()])
        expected_labels = list(page._COLUNAS_LABELS.values())
        assert list(df.columns) == expected_labels

    def test_quantidade_de_linhas(self, page):
        rows = [_make_row(), _make_row(artigo="Artigo B")]
        df = page._rows_to_dataframe(rows)
        assert len(df) == 2

    def test_valores_numericos_preservados(self, page):
        df = page._rows_to_dataframe([_make_row(m2=100.0, kg_m2=1.20)])
        assert df[page._COLUNAS_LABELS["m2"]].iloc[0] == pytest.approx(100.0)
        assert df[page._COLUNAS_LABELS["kg_m2"]].iloc[0] == pytest.approx(1.20)

    def test_none_preservado_como_none(self, page):
        df = page._rows_to_dataframe([_make_row_sem_calculo()])
        assert pd.isna(df[page._COLUNAS_LABELS["kg_m2"]].iloc[0])
        assert pd.isna(df[page._COLUNAS_LABELS["pct_diferenca"]].iloc[0])

    def test_dataframe_vazio(self, page):
        df = page._rows_to_dataframe([])
        assert len(df) == 0
        assert list(df.columns) == list(page._COLUNAS_LABELS.values())

    def test_linha_sem_calculo_preserva_dados_oracle(self, page):
        row = _make_row_sem_calculo()
        df = page._rows_to_dataframe([row])
        assert df[page._COLUNAS_LABELS["artigo"]].iloc[0] == "Artigo B"
        assert df[page._COLUNAS_LABELS["codigo_artigo"]].iloc[0] == "ART999"


# ---------------------------------------------------------------------------
# _exportar_excel
# ---------------------------------------------------------------------------

class TestExportarExcel:
    def test_retorna_bytes(self, page):
        df = page._rows_to_dataframe([_make_row()])
        result = page._exportar_excel(df)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_arquivo_legivel_pelo_pandas(self, page):
        df_original = page._rows_to_dataframe([_make_row(), _make_row(artigo="B")])
        xlsx_bytes = page._exportar_excel(df_original)
        df_lido = pd.read_excel(io.BytesIO(xlsx_bytes))
        assert len(df_lido) == 2

    def test_colunas_presentes(self, page):
        df = page._rows_to_dataframe([_make_row()])
        xlsx_bytes = page._exportar_excel(df)
        df_lido = pd.read_excel(io.BytesIO(xlsx_bytes))
        for label in page._COLUNAS_LABELS.values():
            assert label in df_lido.columns

    def test_valores_numericos_como_numero(self, page):
        # pandas lê inteiros sem decimais como int64 e floats como float64
        # o requisito é que sejam numéricos (não string) — PRD Story 1.4 AC6
        df = page._rows_to_dataframe([_make_row(m2=100.0, pct=-3.62)])
        xlsx_bytes = page._exportar_excel(df)
        df_lido = pd.read_excel(io.BytesIO(xlsx_bytes))
        val_m2 = df_lido[page._COLUNAS_LABELS["m2"]].iloc[0]
        val_pct = df_lido[page._COLUNAS_LABELS["pct_diferenca"]].iloc[0]
        assert pd.api.types.is_number(val_m2), f"m2 esperado numérico, obtido: {type(val_m2)}"
        assert pd.api.types.is_number(val_pct), f"pct esperado numérico, obtido: {type(val_pct)}"
        assert float(val_m2) == pytest.approx(100.0)
        assert float(val_pct) == pytest.approx(-3.62)

    def test_none_exportado_como_nan(self, page):
        df = page._rows_to_dataframe([_make_row_sem_calculo()])
        xlsx_bytes = page._exportar_excel(df)
        df_lido = pd.read_excel(io.BytesIO(xlsx_bytes))
        assert pd.isna(df_lido[page._COLUNAS_LABELS["kg_m2"]].iloc[0])

    def test_multiplas_linhas_integras(self, page):
        rows = [
            _make_row(artigo="Artigo Alpha", m2=100.0),
            _make_row(artigo="Artigo Beta",  m2=200.0),
            _make_row_sem_calculo(),
        ]
        df = page._rows_to_dataframe(rows)
        xlsx_bytes = page._exportar_excel(df)
        df_lido = pd.read_excel(io.BytesIO(xlsx_bytes))
        assert len(df_lido) == 3
        artigos = df_lido[page._COLUNAS_LABELS["artigo"]].tolist()
        assert artigos == ["Artigo Alpha", "Artigo Beta", "Artigo B"]

    def test_nome_aba_correto(self, page):
        from openpyxl import load_workbook
        df = page._rows_to_dataframe([_make_row()])
        xlsx_bytes = page._exportar_excel(df)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Relatório" in wb.sheetnames


# ---------------------------------------------------------------------------
# Constantes de configuração
# ---------------------------------------------------------------------------

class TestConstantes:
    def test_ordem_colunas_tem_12_itens(self, page):
        assert len(page._COLUNAS_ORDEM) == 12

    def test_labels_cobrem_todas_as_colunas(self, page):
        for col in page._COLUNAS_ORDEM:
            assert col in page._COLUNAS_LABELS

    def test_colunas_numericas_sao_subconjunto_da_ordem(self, page):
        assert page._COLUNAS_NUMERICAS.issubset(set(page._COLUNAS_ORDEM))
